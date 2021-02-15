"""Import functionality"""

import logging
import sys
from datetime import datetime
from typing import Iterable, List, Optional

import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from cg.exc import CgError
from cg.store import Store, models

from .models import ApplicationSchema, ApplicationVersionSchema, ParsedApplicationVersion

LOG = logging.getLogger(__name__)


def import_application_versions(
    store: Store, excel_path: str, sign: str, dry_run: bool, skip_missing: bool
):
    """
    Imports all application versions from the specified excel file
    Args:
        :param store:               Status database store
        :param excel_path:          Path to excel file with headers: 'App tag', 'Valid from',
                                    'Standard', 'Priority', 'Express', 'Research'
        :param sign:                Signature of user running the script
        :param dry_run:             Test run, no changes to the database
        :param skip_missing:        Continue despite missing applications
    """
    application_versions: Iterable[ParsedApplicationVersion] = parse_application_versions(
        excel_path=excel_path
    )

    for application_version in application_versions:

        application_obj: models.Application = store.application(application_version.app_tag)

        if not application_obj:
            LOG.error(
                "Failed to find application! Please manually add application: %s",
                application_version.app_tag,
            )

            if skip_missing:
                continue

            LOG.error("Rolling back transaction.")
            store.rollback()
            sys.exit()

        app_tag: str = application_obj.tag
        latest_version: models.ApplicationVersion = store.latest_version(
            application_version.app_tag
        )

        if latest_version and versions_are_same(
            version_obj=latest_version, application_version=application_version
        ):
            LOG.info("skipping redundant application version for app tag %s", app_tag)
            continue

        LOG.info("adding new application version to transaction for app tag %s", app_tag)
        new_version = add_application_version(
            application_obj=application_obj,
            latest_version=latest_version,
            version=application_version,
            sign=sign,
            store=store,
        )
        store.add(new_version)

    if not dry_run:
        LOG.info(
            "all application versions successfully added to transaction, committing " "transaction"
        )
        store.commit()
    else:
        LOG.error("Dry-run, rolling back transaction.")
        store.rollback()


def import_applications(
    store: Store, excel_path: str, sign: str, dry_run: bool, sheet_name: Optional[str] = None
):
    """
    Imports all applications from the specified excel file
    Args:
        :param sheet_name:          Optional sheet name
        :param store:               status database store
        :param excel_path:          Path to excel file with headers: 'App tag', 'Valid from',
                                    'Standard', 'Priority', 'Express', 'Research'
        :param sign:                Signature of user running the script
        :param dry_run:             Test run, no changes to the database
    """
    applications: Iterable[ApplicationSchema] = parse_applications(
        excel_path=excel_path, sheet_name=sheet_name
    )

    for application in applications:
        application_obj: models.Application = store.application(application.tag)
        if application_obj and applications_are_same(
            application_obj=application_obj, application=application
        ):
            LOG.info("skipping redundant application %s", application.tag)
            continue

        LOG.info("adding new application to transaction %s", application.tag)
        new_application = add_application_object(application=application, sign=sign, store=store)
        store.add(new_application)

    if not dry_run:
        LOG.info("all applications successfully added to transaction, committing transaction")
        store.commit()
    else:
        LOG.error("Dry-run, rolling back transaction.")
        store.rollback()


def prices_are_same(first_price: float, second_price: float) -> bool:
    """Checks if the given prices are to be considered equal"""

    if first_price == second_price:
        return True
    if first_price and second_price:
        return int(round(float(first_price))) == int(round(float(second_price)))

    return False


def versions_are_same(
    version_obj: models.ApplicationVersion, application_version: ApplicationVersionSchema
) -> bool:
    """Checks if the given versions are to be considered equal"""
    return (
        version_obj.application.tag == application_version.app_tag
        and version_obj.valid_from == application_version.valid_from
        and prices_are_same(version_obj.price_standard, application_version.standard)
        and prices_are_same(version_obj.price_priority, application_version.priority)
        and prices_are_same(version_obj.price_express, application_version.express)
        and prices_are_same(version_obj.price_research, application_version.research)
    )


def applications_are_same(
    application_obj: models.Application, application: ApplicationSchema
) -> bool:
    """Checks if the given applications are to be considered equal"""

    return application_obj.tag == application.tag


def add_application_version(
    application_obj: models.Application,
    latest_version: Optional[models.ApplicationVersion],
    version: ApplicationVersionSchema,
    sign: str,
    store: Store,
) -> models.ApplicationVersion:
    new_version = store.add_version(
        application=application_obj,
        version=latest_version.version + 1 if latest_version else 1,
        valid_from=version.valid_from,
        prices={
            "standard": version.standard,
            "priority": version.priority,
            "express": version.express,
            "research": version.research,
        },
        comment="Added by %s" % sign,
    )
    return new_version


def add_application_object(
    application: ApplicationSchema, sign: str, store: Store
) -> models.Application:
    """Adds an application from a raw application record"""
    new_application: models.Application = store.add_application(
        tag=application.tag,
        category=application.prep_category,
        description=application.description,
        is_accredited=application.is_accredited == 1.0,
        turnaround_time=application.turnaround_time,
        minimum_order=application.minimum_order,
        sequencing_depth=application.sequencing_depth,
        target_reads=application.target_reads,
        sample_amount=application.sample_amount,
        sample_volume=application.sample_volume,
        sample_concentration=application.sample_concentration,
        priority_processing=application.priority_processing == 1.0,
        details=application.details,
        limitations=application.limitations,
        percent_kth=application.percent_kth,
        percent_reads_guaranteed=application.percent_reads_guaranteed,
        comment=application.comment + " Added by %s" % sign,
        is_archived=application.is_archived == 1.0,
        is_external=application.is_external == 1.0,
        created_at=datetime.now(),
    )
    return new_application


def import_apptags(
    store: Store,
    excel_path: str,
    prep_category: str,
    sign: str,
    sheet_name: str,
    tag_column: int,
    activate: bool,
    inactivate: bool,
):
    """Syncs all applications from the specified excel file"""

    orderform_application_tags = []

    for tag in get_cells_from_excel(
        excel_path=excel_path, sheet_name=sheet_name, tag_column=tag_column
    ):
        LOG.info("Found: %s in orderform", tag)
        orderform_application_tags.append(tag)

    if not orderform_application_tags:
        message = f"No applications found in column {tag_column} (zero-based), exiting"
        raise CgError(message)

    for orderform_application_tag in orderform_application_tags:
        application_obj = store.application(tag=orderform_application_tag)

        if not application_obj:
            message = f"Application {orderform_application_tag} was not found"
            raise CgError(message)

        if application_obj.prep_category != prep_category:
            message = (
                f"{orderform_application_tag} prep_category, expected: {prep_category} was:"
                f" {application_obj.prep_category}"
            )
            raise CgError(message)

        if application_obj.is_archived:
            if activate:
                application_obj.comment = (
                    f"{application_obj.comment}\n{str(datetime.now())[:-10]}"
                    f"Application un-archived by {sign}"
                )
                application_obj.is_archived = False
                LOG.info("Un-archiving %s", application_obj)
            else:
                LOG.warning(
                    "%s is marked as archived but is used in the orderform, consider "
                    "activating it",
                    application_obj,
                )
        else:
            LOG.info("%s is already active, no need to activate it", application_obj)

    all_active_apps_for_category = store.applications(category=prep_category, archived=False)

    for active_application in all_active_apps_for_category:
        if active_application.tag in orderform_application_tags:
            LOG.info("%s was found in orderform tags, no need to archive it", active_application)
            continue
        if not inactivate:
            LOG.warning(
                "%s is marked as active but is not used in the orderform, consider archiving it",
                active_application,
            )
            continue
        active_application.is_archived = True
        active_application.comment = (
            f"{active_application.comment}"
            f"\n{str(datetime.now())[:-10]} "
            f"Application archived by {sign}"
        )
        LOG.info("Archiving %s", active_application)

    if not activate and not inactivate:
        LOG.info("no change mode requested, rolling back transaction")
        store.rollback()
    else:
        LOG.info("all applications successfully synced, committing transaction")
        store.commit()


def get_cells_from_excel(
    excel_path: str, sheet_name: str = "Drop down list", tag_column: int = 2
) -> Iterable[str]:
    workbook: Workbook = openpyxl.load_workbook(filename=excel_path, read_only=True, data_only=True)
    data_sheet: Worksheet = workbook[sheet_name]

    for row in data_sheet.rows:
        value = row[tag_column].value
        if value is None:
            continue
        yield value


def parse_application_versions(
    excel_path: str, sheet_name: Optional[str] = None
) -> Iterable[ParsedApplicationVersion]:
    workbook: Workbook = openpyxl.load_workbook(filename=excel_path, read_only=True, data_only=True)
    data_sheet = workbook.worksheets[0]
    if sheet_name:
        data_sheet = workbook[sheet_name]

    application_info: dict
    for application_info in get_raw_dicts_from_xlsheet(data_sheet):
        yield ParsedApplicationVersion(**application_info)


def parse_applications(
    excel_path: str, sheet_name: Optional[str] = None
) -> Iterable[ApplicationSchema]:
    workbook: Workbook = openpyxl.load_workbook(filename=excel_path, read_only=True, data_only=True)
    data_sheet = workbook.worksheets[0]
    if sheet_name:
        data_sheet = workbook[sheet_name]

    application_info: dict
    for application_info in get_raw_dicts_from_xlsheet(data_sheet):
        yield ApplicationSchema(**application_info)


def get_raw_dicts_from_xlsheet(sheet: Worksheet) -> List[dict]:
    """Get the relevant rows from a sheet."""
    raw_data = []
    header_row = []
    first_row = True

    for row in sheet.rows:
        if row[0].value is None:
            break

        if first_row:
            header_row = [cell.value for cell in row if cell.value is not None]
            first_row = False
        else:
            values = [cell.value for cell in row]
            version_dict = dict(zip(header_row, values))
            raw_data.append(version_dict)

    return raw_data
