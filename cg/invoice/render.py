# -*- coding: utf-8 -*-
import datetime
from pkg_resources import resource_filename

from openpyxl import load_workbook
from openpyxl.styles import Border, Font, PatternFill, Side


def render_xlsx(data, costcenter):
    """Render an Excel invoice."""
    pkg_dir = __name__.rpartition('.')[0]
    template_path = resource_filename(pkg_dir, 'templates/invoice.xlsx')
    wb = load_workbook(template_path)
    ws = wb.active
    ws['C1'] = costcenter.upper()
    ws['F2'] = datetime.datetime.today().date()
    ws['C7'] = data['project']
    ws['C13'] = data['contact']['name']
    ws['C14'] = data['contact']['email']
    ws['C15'] = data['reference']
    ws['C16'] = data['name']
    ws['C17'] = data['address']
    ws['C20'] = data['agreement']

    samples_start = 24
    for index, sample_data in enumerate(data['samples']):
        row = samples_start + index
        ws["A{}".format(row)] = sample_data['name']
        ws["B{}".format(row)] = sample_data['lims_id']
        ws["C{}".format(row)] = sample_data['application_tag']
        ws["D{}".format(row)] = sample_data['project']
        ws["E{}".format(row)] = sample_data['date'].date()
        ws["F{}".format(row)] = sample_data['price']

    ws["E{}".format(row + 2)] = 'Total'
    ws["F{}".format(row + 2)] = "=SUM(F{}, F{})".format(samples_start, row)

    header_rows = [5, 12, 19, 23, row + 2]
    for header_row in header_rows:
        for column in ['A', 'B', 'C', 'D', 'E', 'F']:
            cell = ws["{}{}".format(column, header_row)]
            cell.font = Font(bold=True)
            cell.border = Border(top=Side(border_style='thin', color='000000'),
                                 bottom=Side(border_style='thin', color='000000'))
            cell.fill = PatternFill('solid', fgColor='E5E8E8')

    return wb
