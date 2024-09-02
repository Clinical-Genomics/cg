import datetime as dt

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Font, PatternFill, Side
from pkg_resources import resource_filename


def render_xlsx(data: dict) -> Workbook:
    """Render an Excel invoice template.

    Data is intended to look as follows:

        {
            'invoice_id': str,
            'costcenter': str,
            'project_number': str,
            'customer_id': str,
            'customer_name': str,
            'agreement': str,
            'contact': {
                'name': str,
                'email': str,
                'reference': str,
                'customer_name': str,
                'address': str
            },
            'samples': [{
                'name': str,
         #       'lims_id': str,
                'application_tag': str,
                'project': str,
                'date': str,
                'price': int
            }]
        }
    """
    pkg_dir = __name__.rpartition(".")[0]
    sample_type = "pool" if data["pooled_samples"] else "sample"
    costcenter = data["cost_center"]
    template_path = resource_filename(pkg_dir, f"templates/{costcenter}_{sample_type}_invoice.xlsx")
    workbook = load_workbook(template_path)
    if data["pooled_samples"]:
        worksheet = workbook["Bilaga Prover"]
        worksheet["C1"] = costcenter.upper()
        worksheet["F1"] = f"{data['invoice_id']}-{costcenter}"
        worksheet["F2"] = dt.datetime.today().date()

        samples_start = 7
        for index, lims_sample in enumerate(data["pooled_samples"]):
            total_reads = lims_sample.udf.get("Total Reads (M)", "")
            pool_name = lims_sample.udf.get("pool name", "")

            row = samples_start + index
            worksheet[f"A{row}"] = lims_sample.name
            worksheet[f"B{row}"] = lims_sample.id
            worksheet[f"C{row}"] = total_reads
            worksheet[f"D{row}"] = str(pool_name)

        for column in "ABCDEFG":
            cell = worksheet[f"{column}{6}"]
            cell.font = Font(bold=True)
            cell.border = Border(
                top=Side(border_style="thin", color="000000"),
                bottom=Side(border_style="thin", color="000000"),
            )
            cell.fill = PatternFill("solid", fgColor="E5E8E8")

    worksheet = workbook["Faktura"]
    worksheet.font = Font(size=14)
    for row_dimension in worksheet.row_dimensions.values():
        row_dimension.height = 30
    for col_dimension in worksheet.column_dimensions.values():
        if col_dimension.index in "GHI":
            col_dimension.width = 10
        else:
            col_dimension.width = 18

    worksheet["C1"] = costcenter.upper()
    worksheet["F1"] = f"{data['invoice_id']}-{costcenter}"
    worksheet["F2"] = dt.datetime.today().date()
    worksheet["C7"] = data["project_number"]
    worksheet["C13"] = data["contact"]["name"]
    worksheet["C14"] = data["contact"]["email"]
    worksheet["C15"] = data["contact"]["reference"]
    worksheet["C16"] = data["contact"]["customer_name"]
    worksheet["C17"] = data["contact"]["address"]
    worksheet["C20"] = f"{data['customer_id']} / {data['customer_name']}"

    if data.get("agreement"):
        worksheet["A21"] = "Avtaltets diarienummer"
        worksheet["C21"] = data["agreement"]

    samples_start = 24

    for index, record_data in enumerate(data["records"]):
        row = samples_start + index
        worksheet[f"A{row}"] = record_data["name"]
        worksheet[f"B{row}"] = record_data["lims_id"]
        worksheet[f"C{row}"] = record_data["application_tag"]
        worksheet[f"D{row}"] = record_data["priority"]
        worksheet[f"E{row}"] = record_data["project"]
        worksheet[f"F{row}"] = record_data["date"]
        worksheet[f"G{row}"] = round(record_data["price"])
        if costcenter == "KI":
            worksheet[f"H{row}"] = round(record_data["price_kth"])
            worksheet[f"I{row}"] = round(record_data["total_price"])
    worksheet[f"F{row + 2}"] = "Total:"
    worksheet[f"G{row + 2}"] = f"=SUM(G{samples_start}: G{row})"
    if costcenter == "KI":
        worksheet[f"H{row + 2}"] = f"=SUM(H{samples_start}: H{row})"
        worksheet[f"I{row + 2}"] = f"=SUM(I{samples_start}: I{row})"

    header_rows = [5, 12, 19, 23, row + 2]
    used_columns = "ABCDEFGHI" if costcenter == "KI" else "ABCDEFG"
    for header_row in header_rows:
        for column in used_columns:
            cell = worksheet[f"{column}{header_row}"]
            cell.font = Font(bold=True, size=14)
            cell.border = Border(
                top=Side(border_style="thin", color="000000"),
                bottom=Side(border_style="thin", color="000000"),
            )
            cell.fill = PatternFill("solid", fgColor="E5E8E8")

    return workbook
