import os
from pathlib import Path
from typing import Dict, List, Set

import openpyxl
from cg.constants import ANALYSIS_SOURCES, METAGENOME_SOURCES, DataDelivery, Pipeline
from cg.exc import OrderFormError
from cg.meta.orders import OrderType
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

SEX_MAP = {"male": "M", "female": "F", "unknown": "unknown"}
REV_SEX_MAP = {value: key for key, value in SEX_MAP.items()}
CONTAINER_TYPES = ["Tube", "96 well plate"]
SOURCE_TYPES = set().union(METAGENOME_SOURCES, ANALYSIS_SOURCES)
VALID_ORDERFORMS = [
    "1508:22",  # Orderform MIP, Balsamic, sequencing only, MIP RNA
    "1541:6",  # Orderform Externally sequenced samples
    "1603:9",  # Microbial WGS
    "1604:10",  # Orderform Ready made libraries (RML)
    "1605:8",  # Microbial metagenomes
]
NO_VALUE = "no_value"
NO_ANALYSIS = "no-analysis"

CASE_PROJECT_TYPES = [
    str(OrderType.MIP_DNA),
    str(OrderType.EXTERNAL),
    str(OrderType.BALSAMIC),
    str(OrderType.MIP_RNA),
]


def check_orderform_version(document_title: str) -> None:
    """Raise an error if the orderform is too new or too old for the order portal."""
    for valid_orderform in VALID_ORDERFORMS:
        if valid_orderform in document_title:
            return
    raise OrderFormError(f"Unsupported orderform: {document_title}")


def parse_orderform(excel_path: str) -> dict:
    """Parse out information from an order form."""

    workbook: Workbook = openpyxl.load_workbook(filename=excel_path, read_only=True, data_only=True)

    sheet_name = None
    sheet_names: List[str] = workbook.sheetnames
    for name in ["Orderform", "orderform", "order form"]:
        if name in sheet_names:
            sheet_name = name
            break
    if sheet_name is None:
        raise OrderFormError("'orderform' sheet not found in Excel file")

    orderform_sheet: Worksheet = workbook[sheet_name]
    document_title: str = get_document_title(workbook, orderform_sheet)
    check_orderform_version(document_title)

    raw_samples = relevant_rows(orderform_sheet)

    if len(raw_samples) == 0:
        raise OrderFormError("orderform doesn't contain any samples")
    parsed_samples = [parse_sample(raw_sample) for raw_sample in raw_samples]

    project_type = get_project_type(document_title, parsed_samples)
    delivery_type = get_data_delivery(parsed_samples, OrderType(project_type))

    if project_type in CASE_PROJECT_TYPES:
        parsed_cases = group_cases(parsed_samples)
        items = []
        customer_ids = set()
        for case_id, parsed_case in parsed_cases.items():
            customer_id, case_data = expand_case(case_id, parsed_case, delivery_type)
            customer_ids.add(customer_id)
            items.append(case_data)
    else:
        customer_ids = set(sample["customer"] for sample in parsed_samples)
        items = parsed_samples

    customer_options = len(customer_ids)
    if customer_options == 0:
        raise OrderFormError("Customer information is missing")
    elif customer_options != 1:
        raise OrderFormError(f"Samples have different customers: {customer_ids}")

    filename_base = Path(excel_path).stem
    parsed_order = {
        "customer": customer_ids.pop(),
        "delivery_type": delivery_type,
        "items": items,
        "name": filename_base,
        "project_type": project_type,
    }

    return parsed_order


def get_document_title(workbook: Workbook, orderform_sheet: Worksheet) -> str:
    """Get the document title for the order form.

    Openpyxl use 1 based counting
    """
    for sheet_number, sheet_name in enumerate(workbook.sheetnames):
        if sheet_name.lower() == "information":
            information_sheet: Worksheet = workbook[sheet_name]
            document_title = information_sheet.cell(1, 3).value
            return document_title

    document_title = orderform_sheet.cell(1, 2).value
    return document_title


def get_project_type(document_title: str, parsed_samples: List) -> str:
    """Determine the project type."""

    project_type = None

    if "1541" in document_title:
        project_type = "external"
    elif "1604" in document_title:
        project_type = "rml"
    elif "1603" in document_title:
        project_type = "microsalt"
    elif "1605" in document_title:
        project_type = "metagenome"
    elif "1508" in document_title:
        analyses = set(sample["data_analysis"] for sample in parsed_samples)

        if len(analyses) != 1:
            raise OrderFormError(f"mixed 'Data Analysis' types: {', '.join(analyses)}")

        analysis = analyses.pop().lower().replace(" ", "-")

        if analysis == NO_ANALYSIS:
            project_type = "fastq"
        else:
            project_type = analysis

    return project_type


def _parse_data_analysis(samples) -> str:
    data_analyses = set()
    for sample in samples:
        data_analyses.add(sample["data_analysis"])

    if len(data_analyses) > 1:
        raise OrderFormError(f"mixed 'Data Analysis' types: {', '.join(data_analyses)}")

    return data_analyses.pop().lower().replace(" ", "-")


def _is_from_orderform_without_data_delivery(data_delivery: str) -> bool:
    return data_delivery == NO_VALUE


def get_data_delivery(samples: [dict], project_type: OrderType) -> DataDelivery:
    """Determine the order_data delivery type."""

    data_delivery: str = _parse_data_delivery(samples)

    if _is_from_orderform_without_data_delivery(data_delivery):

        if project_type == OrderType.FLUFFY:
            return DataDelivery.NIPT_VIEWER

        if project_type == OrderType.METAGENOME:
            return DataDelivery.FASTQ

        if project_type == OrderType.MICROSALT:
            data_analysis: str = _parse_data_analysis(samples)

            if data_analysis == "fastq":
                return DataDelivery.FASTQ
            if data_analysis == "custom":
                return DataDelivery.FASTQ_QC

        if project_type == OrderType.RML:
            return DataDelivery.FASTQ

        if project_type == OrderType.EXTERNAL:
            return DataDelivery.SCOUT

        raise OrderFormError(f"Could not determine value for Data Delivery")

    if data_delivery == "analysis-+-bam":
        return DataDelivery.ANALYSIS_BAM_FILES

    try:
        return DataDelivery(data_delivery)
    except ValueError:
        raise OrderFormError(f"Unsupported Data Delivery: {data_delivery}")


def _parse_data_delivery(samples: [dict]) -> str:

    data_deliveries = set()
    for sample in samples:
        data_delivery = sample.get("data_delivery", NO_VALUE) or NO_VALUE
        data_deliveries.add(data_delivery)

    if len(data_deliveries) > 1:
        raise OrderFormError(f"mixed 'Data Delivery' types: {', '.join(data_deliveries)}")

    return data_deliveries.pop().lower().replace(" ", "-")


def expand_case(case_id: str, parsed_case: dict, data_delivery: DataDelivery) -> tuple:
    """Fill-in information about case."""
    new_case = {"name": case_id, "samples": []}
    samples = parsed_case["samples"]

    require_qcoks = set(raw_sample["require_qcok"] for raw_sample in samples)
    new_case["require_qcok"] = True in require_qcoks

    priorities = set(raw_sample["priority"] for raw_sample in samples)

    if len(priorities) != 1:
        raise OrderFormError(f"multiple values for 'Priority' for case: {case_id}")

    new_case["priority"] = priorities.pop()

    customers = set(raw_sample["customer"] for raw_sample in samples)
    if len(customers) != 1:
        raise OrderFormError("Invalid customer information: {}".format(customers))
    customer = customers.pop()
    gene_panels = set()
    for raw_sample in samples:
        if raw_sample["panels"]:
            gene_panels.update(raw_sample["panels"])
        new_sample = {
            "name": raw_sample["name"],
            "sex": raw_sample["sex"],
            "application": raw_sample["application"],
            "source": raw_sample["source"],
            "data_delivery": str(data_delivery),
        }
        if raw_sample.get("container") in CONTAINER_TYPES:
            new_sample["container"] = raw_sample["container"]

        for key in (
            "capture_kit",
            "comment",
            "container_name",
            "data_analysis",
            "elution_buffer",
            "formalin_fixation_time",
            "from_sample",
            "post_formalin_fixation_time",
            "quantity",
            "status",
            "time_point",
            "tissue_block_size",
            "tumour",
            "tumour_purity",
            "volume",
            "well_position",
        ):
            if raw_sample.get(key):
                new_sample[key] = raw_sample[key]

        for parent_id in ("mother", "father"):
            if raw_sample[parent_id]:
                new_sample[parent_id] = raw_sample[parent_id]
        new_case["samples"].append(new_sample)

    new_case["panels"] = list(gene_panels)

    return customer, new_case


def group_cases(parsed_samples):
    """Group samples on case."""
    raw_cases = {}
    for sample in parsed_samples:
        case_id = sample["case"]
        if case_id not in raw_cases:
            raw_cases[case_id] = {"samples": []}
        raw_cases[case_id]["samples"].append(sample)
    return raw_cases


def parse_sample(raw_sample: Dict[str, str]) -> dict:
    """Parse a raw sample row from order form sheet."""
    if ":" in raw_sample.get("UDF/Gene List", ""):
        raw_sample["UDF/Gene List"] = raw_sample["UDF/Gene List"].replace(":", ";")

    if raw_sample["UDF/priority"].lower() == "förtur":
        raw_sample["UDF/priority"] = "priority"
    raw_source = raw_sample.get("UDF/Source")
    sample = {
        "application": raw_sample["UDF/Sequencing Analysis"],
        "capture_kit": raw_sample.get("UDF/Capture Library version"),
        "case": raw_sample.get("UDF/familyID"),
        "comment": raw_sample.get("UDF/Comment"),
        "container": raw_sample.get("Container/Type"),
        "container_name": raw_sample.get("Container/Name"),
        "custom_index": raw_sample.get("UDF/Custom index"),
        "customer": raw_sample["UDF/customer"],
        "data_analysis": raw_sample["UDF/Data Analysis"],
        "data_delivery": raw_sample.get("UDF/Data Delivery"),
        "elution_buffer": raw_sample.get("UDF/Sample Buffer"),
        "extraction_method": raw_sample.get("UDF/Extraction method"),
        "formalin_fixation_time": raw_sample.get("UDF/Formalin Fixation Time"),
        "index": raw_sample.get("UDF/Index type"),
        "from_sample": raw_sample.get("UDF/is_for_sample"),
        "name": raw_sample["Sample/Name"],
        "organism": raw_sample.get("UDF/Strain"),
        "organism_other": raw_sample.get("UDF/Other species"),
        "panels": (
            raw_sample["UDF/Gene List"].split(";") if raw_sample.get("UDF/Gene List") else None
        ),
        "pool": raw_sample.get("UDF/pool name"),
        "post_formalin_fixation_time": raw_sample.get("UDF/Post Formalin Fixation Time"),
        "priority": raw_sample["UDF/priority"].lower() if raw_sample.get("UDF/priority") else None,
        "reagent_label": raw_sample.get("Sample/Reagent Label"),
        "reference_genome": raw_sample.get("UDF/Reference Genome Microbial"),
        "require_qcok": raw_sample.get("UDF/Process only if QC OK") == "yes",
        "rml_plate_name": raw_sample.get("UDF/RML plate name"),
        "sex": REV_SEX_MAP.get(raw_sample.get("UDF/Gender", "").strip()),
        "source": raw_source if raw_source in SOURCE_TYPES else None,
        "status": raw_sample["UDF/Status"].lower() if raw_sample.get("UDF/Status") else None,
        "tissue_block_size": raw_sample.get("UDF/Tissue Block Size"),
        "tumour": raw_sample.get("UDF/tumor") == "yes",
        "tumour_purity": raw_sample.get("UDF/tumour purity"),
        "well_position": raw_sample.get("Sample/Well Location"),
        "well_position_rml": raw_sample.get("UDF/RML well position"),
    }

    numeric_attributes = [
        ("index_number", "UDF/Index number"),
        ("volume", "UDF/Volume (uL)"),
        ("quantity", "UDF/Quantity"),
        ("concentration", "UDF/Concentration (nM)"),
        ("concentration_sample", "UDF/Sample Conc."),
        ("time_point", "UDF/time_point"),
    ]
    for json_key, excel_key in numeric_attributes:
        str_value = raw_sample.get(excel_key, "").rsplit(".0")[0]
        if str_value.replace(".", "").isnumeric():
            sample[json_key] = str_value

    for parent in ["mother", "father"]:
        parent_key = f"UDF/{parent}ID"
        sample[parent] = (
            raw_sample[parent_key]
            if raw_sample.get(parent_key) and (raw_sample[parent_key] != "0.0")
            else None
        )

    return sample


def relevant_rows(orderform_sheet: Worksheet) -> List[Dict[str, str]]:
    """Get the relevant rows from an order form sheet."""
    raw_samples = []
    current_row = None
    empty_row_found = False
    row: tuple
    for row in orderform_sheet.rows:
        if row[0].value == "</SAMPLE ENTRIES>":
            break

        if current_row == "header":
            header_row = [cell.value for cell in row]
            current_row = None
        elif current_row == "samples":
            values = []
            for cell in row:
                value = str(cell.value)
                if value == "None":
                    value = ""
                values.append(value)

            # skip empty rows
            if values[0]:
                if empty_row_found:
                    raise OrderFormError(
                        f"Found data after empty lines. Please delete any "
                        f"non-sample data rows in between the samples"
                    )

                sample_dict = dict(zip(header_row, values))
                raw_samples.append(sample_dict)
            else:
                empty_row_found = True

        if row[0].value == "<TABLE HEADER>":
            current_row = "header"
        elif row[0].value == "<SAMPLE ENTRIES>":
            current_row = "samples"
    return raw_samples
