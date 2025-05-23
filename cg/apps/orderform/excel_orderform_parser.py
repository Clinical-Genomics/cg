import logging
from pathlib import Path

import openpyxl
from openpyxl.cell.cell import Cell
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import ConfigDict, TypeAdapter

from cg.apps.orderform.orderform_parser import OrderformParser
from cg.apps.orderform.utils import are_all_samples_metagenome
from cg.constants import DataDelivery
from cg.constants.orderforms import Orderform
from cg.exc import OrderFormError
from cg.models.orders.constants import OrderType
from cg.models.orders.excel_sample import ExcelSample

LOG = logging.getLogger(__name__)


class ExcelOrderformParser(OrderformParser):
    model_config = ConfigDict(arbitrary_types_allowed=True)
    NO_ANALYSIS: str = "no-analysis"
    NO_VALUE: str = "no_value"
    SHEET_NAMES: list[str] = ["Orderform", "orderform", "order form", "Order Form"]
    VALID_ORDERFORMS: list[str] = [
        f"{Orderform.MIP_DNA}:{Orderform.get_current_orderform_version(Orderform.MIP_DNA)}",  # Orderform MIP-DNA, Balsamic, sequencing only, MIP-RNA
        f"{Orderform.MICROSALT}:{Orderform.get_current_orderform_version(Orderform.MICROSALT)}",  # Microbial WHOLE_GENOME_SEQUENCING
        f"{Orderform.NALLO}:{Orderform.get_current_orderform_version(Orderform.NALLO)}",  # Nallo
        f"{Orderform.RML}:{Orderform.get_current_orderform_version(Orderform.RML)}",  # Orderform Ready made libraries (RML)
        f"{Orderform.METAGENOME}:{Orderform.get_current_orderform_version(Orderform.METAGENOME)}",  # Microbial meta genomes
        f"{Orderform.SARS_COV_2}:{Orderform.get_current_orderform_version(Orderform.SARS_COV_2)}",  # Orderform SARS-CoV-2
        f"{Orderform.MICROBIAL_FASTQ}:{Orderform.get_current_orderform_version(Orderform.MICROBIAL_FASTQ)}",  # Microbial FASTQ
        f"{Orderform.PACBIO_LONG_READ}:{Orderform.get_current_orderform_version(Orderform.PACBIO_LONG_READ)}",
    ]
    samples: list[ExcelSample] = []

    def check_orderform_version(self, document_title: str) -> None:
        """Raise an error if the orderform is too new or too old for the order portal"""
        LOG.info(f"Validating that {document_title} is a correct orderform version")
        for valid_orderform in self.VALID_ORDERFORMS:
            if valid_orderform in document_title:
                return
        raise OrderFormError(f"Unsupported orderform: {document_title}")

    def get_sheet_name(self, sheet_names: list[str]) -> str:
        """Return the correct (existing) sheet names"""

        for name in sheet_names:
            if name in self.SHEET_NAMES:
                LOG.info(f"Found sheet name {name}")
                return name
        raise OrderFormError("'orderform' sheet not found in Excel file")

    @staticmethod
    def get_document_title(workbook: Workbook, orderform_sheet: Worksheet) -> str:
        """Get the document title for the order form.

        Openpyxl use 1 based counting
        """
        for sheet_number, sheet_name in enumerate(workbook.sheetnames):
            if sheet_name.lower() != "information":
                continue
            information_sheet: Worksheet = workbook[sheet_name]
            document_title = information_sheet.cell(1, 3).value
            LOG.info(f"Found document title {document_title}")
            return document_title

        document_title = orderform_sheet.cell(1, 2).value
        LOG.info(f"Found document title {document_title}")
        return document_title

    @staticmethod
    def get_sample_row_info(
        row: tuple[Cell], header_row: list[str], empty_row_found: bool
    ) -> dict | None:
        """Convert an Excel row with sample data into a dict with sample info"""
        values = []
        cell: Cell
        for cell in row:
            value = str(cell.value)
            if value == "None":
                value = ""
            if value == "NA":
                value = None
            values.append(value)

        if not values[0]:
            return None
        # skip empty rows
        if empty_row_found:
            raise OrderFormError(
                "Found data after empty lines. Please delete any non-sample data rows in between the samples"
            )

        sample_dict = dict(zip(header_row, values))
        sample_dict.pop(None)
        return sample_dict

    @staticmethod
    def get_header(rows: list[tuple[Cell]]) -> list[str]:
        header_row: list[str] = []
        header = False
        for row in rows:
            if header:
                return [cell.value for cell in row]
            if row[0].value == "<TABLE HEADER>":
                LOG.debug("Found header row")
                header = True
        return header_row

    @staticmethod
    def get_raw_samples(rows: list[tuple[Cell]], header_row: list[str]) -> list[dict]:
        raw_samples: list[dict] = []
        sample_rows = False
        empty_row_found = False
        for row in rows:
            if row[0].value == "</SAMPLE ENTRIES>":
                LOG.debug("End of samples info")
                return raw_samples

            if sample_rows:
                sample_dict: dict | None = ExcelOrderformParser.get_sample_row_info(
                    row=row, header_row=header_row, empty_row_found=empty_row_found
                )
                if sample_dict:
                    raw_samples.append(sample_dict)
                else:
                    empty_row_found = True

            elif row[0].value == "<SAMPLE ENTRIES>":
                LOG.debug("Found samples row")
                sample_rows = True
        return raw_samples

    @staticmethod
    def relevant_rows(orderform_sheet: Worksheet) -> list[dict[str, str]]:
        """Get the relevant rows from an order form sheet."""
        # orderform_sheet.rows is a generator. Convert to list to be able to iterate multiple times
        rows = list(orderform_sheet.rows)
        header_row: list[str] = ExcelOrderformParser.get_header(rows)
        return ExcelOrderformParser.get_raw_samples(rows=rows, header_row=header_row)

    def get_project_type(self, document_title: str) -> str:
        """Determine the project type and set it to the class."""
        document_number_to_project_type = {
            Orderform.MICROSALT: OrderType.MICROSALT,
            Orderform.SARS_COV_2: OrderType.SARS_COV_2,
            Orderform.MICROBIAL_FASTQ: OrderType.MICROBIAL_FASTQ,
            Orderform.NALLO: OrderType.NALLO,
            Orderform.PACBIO_LONG_READ: OrderType.PACBIO_LONG_READ,
        }
        for document_number, value in document_number_to_project_type.items():
            if document_number in document_title:
                return value

        analysis = self.parse_data_analysis()
        if Orderform.RML in document_title:
            return OrderType.RML if analysis == self.NO_ANALYSIS else analysis
        if Orderform.MIP_DNA in document_title:
            if analysis == self.NO_ANALYSIS:
                return (
                    OrderType.METAGENOME
                    if are_all_samples_metagenome(self.samples)
                    else OrderType.FASTQ
                )
            else:
                return analysis
        raise OrderFormError(f"Undetermined project type in: {document_title}")

    def parse_data_analysis(self) -> str:
        data_analyses = {sample.data_analysis for sample in self.samples if sample.data_analysis}

        if len(data_analyses) > 1:
            raise OrderFormError(f"mixed 'Data Analysis' types: {', '.join(data_analyses)}")

        return data_analyses.pop().lower().replace(" ", "-")

    def get_data_delivery(self) -> str:
        "Get the data delivery type."
        data_deliveries: set[str] = {
            sample.data_delivery or self.NO_VALUE for sample in self.samples
        }
        if len(data_deliveries) > 1:
            raise OrderFormError(f"mixed 'Data Delivery' types: {', '.join(data_deliveries)}")

        return self._transform_data_delivery(data_deliveries.pop())

    def get_customer_id(self) -> str:
        """Set the customer id"""

        customers = {sample.customer for sample in self.samples}
        if len(customers) != 1:
            raise OrderFormError(f"Invalid customer information: {customers}")
        return customers.pop()

    def parse_orderform(self, excel_path: str) -> None:
        """Parse out information from an order form."""

        LOG.info(f"Open excel workbook from file {excel_path}")
        workbook: Workbook = openpyxl.load_workbook(
            filename=excel_path, read_only=True, data_only=True
        )

        sheet_name: str = self.get_sheet_name(workbook.sheetnames)

        orderform_sheet: Worksheet = workbook[sheet_name]
        document_title: str = self.get_document_title(
            workbook=workbook, orderform_sheet=orderform_sheet
        )
        self.check_orderform_version(document_title)

        LOG.info("Parsing all samples from orderform")
        raw_samples: list[dict] = self.relevant_rows(orderform_sheet)

        if not raw_samples:
            raise OrderFormError("orderform doesn't contain any samples")

        excel_sample_list_validator = TypeAdapter(list[ExcelSample])
        self.samples: list[ExcelSample] = excel_sample_list_validator.validate_python(raw_samples)
        self.project_type: str = self.get_project_type(document_title)
        self.delivery_type = self.get_data_delivery()
        self.customer_id = self.get_customer_id()
        self.order_name = Path(excel_path).stem

    @staticmethod
    def _transform_data_delivery(data_delivery: str) -> str:
        """Transforms the data-delivery parsed in the excel file, to the ones used in cg."""
        try:
            orderform_to_internal: dict = {
                "analysis": DataDelivery.ANALYSIS_FILES,
                "analysis + scout": DataDelivery.ANALYSIS_SCOUT,
                "bam": DataDelivery.BAM,
                "fastq": DataDelivery.FASTQ,
                "fastq + analysis": DataDelivery.FASTQ_ANALYSIS,
                "fastq + analysis + scout": DataDelivery.FASTQ_ANALYSIS_SCOUT,
                "fastq + Scout": DataDelivery.FASTQ_SCOUT,
                "fastq qc": DataDelivery.FASTQ_QC,
                "fastq qc + analysis": DataDelivery.FASTQ_QC_ANALYSIS,
                "no delivery": DataDelivery.NO_DELIVERY,
                "scout": DataDelivery.SCOUT,
                "statina": DataDelivery.STATINA,
                "fastq-analysis": DataDelivery.FASTQ_ANALYSIS,  # Sars Cov10 orderform does not have the same options as others
                "raw_data + analysis": DataDelivery.RAW_DATA_ANALYSIS,
                "raw_data + scout": DataDelivery.RAW_DATA_SCOUT,
                "raw_data + analysis + scout": DataDelivery.RAW_DATA_ANALYSIS_SCOUT,
            }
            return orderform_to_internal[data_delivery]
        except KeyError as error:
            raise OrderFormError(f"Unsupported Data Delivery: {data_delivery}") from error
